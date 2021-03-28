from flask import Flask, render_template, request, redirect, url_for, session, send_file
from flask import *
from werkzeug.utils import secure_filename
import openpyxl
import requests

app = Flask(__name__)
xcel_path = ""

@app.route('/')
def upload():
  return render_template('index.html')

@app.route('/uploader', methods = ['GET', 'POST'])
def upld_file():
    global xcel_path
    if request.method == 'POST':
      cities = request.files['xl']
      list  = str(cities).split("'")
      if list[1] != '':
        cities.save(secure_filename(cities.filename))
        xcel_path = cities.filename
        #print(xcel_path)
        obj = openpyxl.load_workbook(xcel_path)
        sheet = obj.active
        row = sheet.max_row
        for i in range(2, row + 1):
          city = sheet.cell(row = i,column = 1)
          unit = sheet.cell(row = i,column = 2)
          if unit.value == 'F' :
            tempUnit = 'imperial'
          else:
            tempUnit = 'metric'
          url = 'http://api.openweathermap.org/data/2.5/weather?apikey=4a2360d14bf33378079d2e2d49e35ddb&mode=json&units={}&q={}'.format(tempUnit,city.value)
          response = requests.get(url) 
          x = response.json()
          if x['cod'] != '404':
            sheet.cell(row = i,column = 3, value = x['main']['temp'])
            sheet.cell(row = i,column = 4, value = x['main']['humidity'])
            sheet.cell(row = i,column = 5, value = x['main']['pressure'])
            sheet.cell(row = i,column = 6, value = x['weather'][0]['description'])
          else:
            sheet.cell(row = i,column = 3, value = "City not found")

      obj.save('Excel_file.xlsx')
    return render_template('download.html')

@app.route('/about')
def about():
  return render_template('about.html')

@app.route('/sample')
def sample():
  return render_template('sample.html')

@app.route('/download/<int:x>')
def download_file(x):
  global xcel_path
  error = None
  e = 'C:/Users/Hp/Desktop/WeatherApp/Excel file.xlsx'
  d = 'Excel_file.xlsx'
  if x == 1:
      return send_file(e, as_attachment = True)
  if xcel_path == "":
    error = "Please execute an Excel file first"
    return render_template('download.html',error = error)

  else:
    if x == 2:
      return send_file(d, as_attachment = True)


if (__name__=='__main__'):
    app.run(debug=True)
